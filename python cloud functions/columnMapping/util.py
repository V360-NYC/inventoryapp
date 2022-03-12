# D360 Utility Module: This util should be synced accross both code modules
# So no change in function name or static data structure. 
# No direct access to dictionary data-structure.
# Any function that is part of Utility module should have time complexity of O(1).
import re
import logging
import datetime
import validators
import numpy as np
import openpyxl
import pandas as pd
# Gets or creates a logger
logger = logging.getLogger(__name__)  

# set log level
logger.setLevel(logging.DEBUG)

# define file handler and set formatter
date_now = datetime.datetime.now()
file_handler = logging.FileHandler('test'+str(date_now)+'.log')
formatter    = logging.Formatter('%(asctime)s : %(levelname)s : %(name)s : %(message)s')
file_handler.setFormatter(formatter)

# add file handler to logger
logger.addHandler(file_handler)

# Logs
logger.info('Logger Created Successfully!!')
# logger.debug('A debug message')

# logger.warning('Something is not right.')
# logger.error('A Major error has happened.')
# logger.critical('Fatal error. Cannot continue')


# Note: All dictionary should be in upper case. 

actual_heading_dict={
'SHAPE':'SHAPE','SHP':'SHAPE',
'COLOR':'COLOR','COL':'COLOR','COLOUR':'COLOR','CLR':'COLOR',
'CUT GRADE':'CUT','CUT':'CUT',
'CLARITY':'PURITY','CLR':'PURITY','PURITY':'PURITY',
'SYMMETRY':'SYMN','SYM':'SYMN','SYMN':'SYMN',
'POLISH':'POLISH','POL':'POLISH',
'WEIGHT':'WEIGHT',
'TABLE':'TABLE','TABLE %':'TABLE','TBL':'TABLE','TABLE%':'TABLE','TABLEPER':'TABLE',
'REPORTNO':'REPORTNO','REPORT NO':'REPORTNO','REP NO':'REPORTNO','REPORT #':'REPORTNO',
'CERT':'CERT','LAB':'CERT',
'MES1':'MES1','LENGTH':'MES1','MAX':'MES1','MEASUREMENTS':'MES1',
'MES2':'MES2','WIDTH':'MES2','MIN':'MES2',
'MES3':'MES3','HGT':'MES3','DEPTH':'MES3',
'PDFLINK':'PDFLINK',
'CSTATUS':'CSTATUS',
'VIDEOLINK':'VIDEOLINK',
'RATIO':'RATIO',
'TOTAL':'TOTAL','CASH PRICE':'TOTAL','AMT $':'TOTAL','AMOUNT':'TOTAL',
'DEPTH':'DEPTHPER','DEPTH %':'DEPTHPER','DEPTHPER':'DEPTHPER','DEPTH%':'DEPTHPER','TD%':'DEPTHPER',
'FLUORESCENCE INTENSITY':'FLUOR','FLUORESCENCE':'FLUOR','FLUOR':'FLUOR','FLOUR':'FLUOR'}

def skip_to(input_df):
    for index, values in input_df.iterrows():
        for item in values:
            # print(values)
            # if item =="Report #":
            if str(item).strip().upper() in actual_heading_dict:
                # print(item)
                return pd.DataFrame(
                    input_df.iloc[index + 1 : ].to_numpy(), 
                    columns = input_df.iloc[index])
    raise Exception('header row not found')

def getActualHeading(input_heading):
    # skip_to(input_df)
    if input_heading.strip().upper() in actual_heading_dict:
        return actual_heading_dict[input_heading.strip().upper()]
    else:
        logger.warning(f"The given heading: {input_heading} is not available in our dictionary")
        return "NOT_REQUIRED"

master_dict={

'actual_shape_dict': {'ROUND':'RD', 'RD':'RD', 'R':'RD', 'BR':'RD', 'RB':'RD', 'ROUND BRILLIANT':'RD', 'ROUNDBRILLIANT':'RD',     ' BRILLIANT':'RD',
          'BRILLIANT CUT':'RD', 'BRILLIANTCUT':'RD', 
          'OVAL':'OV',  'OV':'OV', 'OC':'OV',  'OVEL':'OV', 'OL':'OV', 
          'EMERALD':'EM', 'EM':'EM', 'EMRD':'EM', 'EC':'EM', 
          'CUSHION MODIFIED':'CU', 'CMB':'CU','CM':'CU', 'CS':'CU', 'CUSHIONMODIFIED':'CU',
          'CUSHION':'CU', 'CUS':'CU','CU':'CU',
          'PRINCESS':'PR','PR':'PR','PC':'PR',
          'PEAR':'PS','PAER':'PS', 'PER':'PS', 'PS':'PS',
          'RADIANT':'RA', 'RAD':'RA', 'RA':'RA',          
          'MARQUISE':'MQ', 'MR':'MQ', 'MQ':'MQ', 'MAR':'MQ',
          'ASHCHER':'AS', 'AS':'AS', 'ASSCHER': 'AS',
          'HEART':'HS','HRT':'HS', 'LOVE':'HS', 'HS':'HS', 'HR':'HS', 'HC':'HS',
          'TRIANGLE':'TR', 'TRI': 'TR', 'TR':'TR'},

'actual_color_dict' : {'D':'D', 'E':'E', 'F':'F', 'D-':'D', 'E-':'E', 'F-':'F','D+':'D', 'E+':'E', 'F+':'F', 'G':'G', 'H':'H',     'I':'I', 'G-':'G', 'H-':'H', 'I-':'I', 'G+':'G', 'H+':'H', 'I+':'I','J':'J', 'K':'K', 'L':'L', 'J-':'J', 'K-':'K', 'L-':'L', 'J+':'J', 'K+':'K', 'L+':'L','M':'M', 'N':'N', 'O':'O', 'M-':'M', 'N-':'N', 'O-':'O', 'M+':'M', 'N+':'N', 'O+':'O','P':'P', 'Q':'Q', 'P-':'P', 'Q-':'Q', 'Q+':'Q', 'P+':'P','Q':'Q', 'R':'R', 'S':'S', 'T':'T', 'U':'U', 'V':'V','W':'W', 'X':'X', 'Y':'Y', 'Z':'Z'},
    
'actual_fluor_dict' :{ 'NONE':'N', 'NON':'N', 'N':'N', 'NO':'N', 'NAN':'N', 'FAINT':'F','FNT':'F', 'FAINT':'F', 'F': 'F',
    'MEDIUM':'M','MED':'M', 'M':'M', 'MEDIUMYELLOW': 'M', 'STRONG':'S', 'STG':'S', 'S':'S', 'ST':'S', 'STRONGYELLOW':'S',
    'VERY STRONG':'VS', 'VST':'VS', 'VSTG':'VS', 'VS':'VS', 'VERYSTRONG':'VS', 'VERYSTRONGBL': 'VS',
    },

'actual_purity_dict' : { 'FL':'FL', 'IF':'IF', 'VVS1':'VVS1', 'VVS2':'VVS2', 'VS1':'VS1',
                    'FL-':'IF', 'IF-':'IF', 'VVS1-':'VVS1', 'VVS2-':'VVS2', 'VS1-':'VS1',
                    'FL+':'IF', 'IF+':'IF', 'VVS1+':'VVS1', 'VVS2+':'VVS2', 'VS1+':'VS1',
                    'VS2':'VS2', 'SI1':'SI1', 'SI2':'SI2', 'I1':'I1', 'I2':'I2', 'I3':'I3',
                    'VS2-':'VS2', 'SI1-':'SI1', 'SI2-':'SI2', 'I1-':'I1', 'I2-':'I2', 'I3-':'I3',
                    'VS2+':'VS2', 'SI1+':'SI1', 'SI2+':'SI2', 'I1+':'I1', 'I2+':'I2', 'I3+':'I3'},


'actual_cut_dict' : {'EX':'X', 'VG':'VG', 'G': 'G', 'X':'X','EX-':'X', 'VG-':'VG', 'G-': 'G','EX+':'X', 'VG+':'VG', 'G+': 'G',
                   'EXCELLENT':'X', 'VERY GOOD': 'VG', 'VERYGOOD': 'VG', 'GOOD': 'G','F':'F','FAIR':'F','P':'P','POOR':'P'},

'actual_polish_dict' : {'EX':'X', 'VG':'VG', 'G': 'G', 'X':'X',
                   'EX-':'X', 'VG-':'VG', 'G-': 'G',
                   'EX+':'X', 'VG+':'VG', 'G+': 'G',
                   'EXCELLENT':'X', 'VERY GOOD': 'VG', 'VERYGOOD': 'VG', 'GOOD': 'G',
                   'F':'F','FAIR':'F','P':'P','POOR':'P',},

'actual_symn_dict' : {'EX':'X', 'VG':'VG', 'G': 'G', 'X':'X',
                   'EX-':'X', 'VG-':'VG', 'G-': 'G',
                   'EX+':'X', 'VG+':'VG', 'G+': 'G',
                   'EXCELLENT':'X', 'VERY GOOD': 'VG', 'VERYGOOD': 'VG', 'GOOD': 'G',
                   'F':'F','FAIR':'F','P':'P','POOR':'P',},
}

headings_required=["SHAPE","COLOR","FLUOR","CUT","SYMN","POLISH","PURITY"]
def getActualValue(input_value_str,input_value_df,filePath):
    if input_value_str in headings_required:
        input_dict="actual_"+input_value_str.lower()+"_dict"
        actual_dict=master_dict[input_dict] 
        input_value_df.replace(actual_dict,inplace=True)
        actual_values = (list(actual_dict.values()))
        actual_values.append(np.NaN)
        keyerror_values = set(list(input_value_df[~input_value_df.isin(actual_values)]))
        # keyerror_values
        if keyerror_values:
            logger.warning(f"The given values {keyerror_values} is not available in our {input_value_str} dictionary")
    elif input_value_str in ["PDFLINK","VIDEOLINK"]:
        input_value_df=getActualLink(filePath,input_value_str,input_value_df) 
    return input_value_df


def url_ok(url):
    return validators.url(url)

def getActualLink(filePath,input_value_str,input_value_df):
    wb = openpyxl.load_workbook(filePath)
    ws = wb.worksheets[0]
    total_rows=len(input_value_df.index)
    
    total_columns=ws.max_column
    for heading in range(1,total_columns):
        print(ws.cell(1,heading).value)
        if ws.cell(1,heading).value.upper() == input_value_str.upper():
            print(input_value_str,ws.cell(1,heading).value)
            index=heading

    ls=[]
    count=0
    print(ws.cell(total_rows+1,index).value)
    for row in range(2,total_rows+2):
        count+=1
        if ws.cell(row,index).value:
            if ws.cell(row,index).hyperlink:
                if not url_ok(ws.cell(row,index).hyperlink.target):
                    logger.warning(f"The given hyperlink url: {ws.cell(row,index).hyperlink.target} is not valid, in this column {input_value_str} and on the row number {row}")
                    ls.append(f'Link not working:{ws.cell(row,index).hyperlink.target}')
                else:
                    ls.append(ws.cell(row,index).hyperlink.target)
            else:
                if not url_ok(ws.cell(row,index).value):
                    logger.warning(f"The given url: {ws.cell(row,index).value} is not valid, in this column {input_value_str} and on row number {row}")
                    ls.append(f'Link not working:{ws.cell(row,index).value}')
                else:
                    ls.append(ws.cell(row,index).value)
        else:
            ls.append('')
    print(f'total {total_rows} len {len(ls)} count {count}')
    return ls

def getActualMeasurement(input_measurement_df):
    input_measurement_df=input_measurement_df.str.replace('[a-zA-Z]', '-', regex=True)
    result=input_measurement_df.str.split("-",expand=True)
    return result[0],result[1],result[2]




        


# # It will generate key by joining all properties by ',' and return it.
# def dict_key(weight, shape, color, clarity, fluor, cut, polish, sym): #WSCCFCPS
#   return ','.join([get_d360_weight(weight), shape, color, clarity, fluor, cut, polish, sym]).strip().upper()


# def user_value_dict_key(weight, shape, color, clarity, fluor, cut, polish, sym) :
#   weight = weight
#   shape = actual_shape_dict[shape.strip().upper()]
#   color = actual_color_dict[color.strip().upper()]
#   clarity = actual_clarity_dict[clarity.strip().upper()]
#   fluor = actual_fluor_dict[fluor.strip().upper()]
#   cut = actual_cut_dict[cut.strip().upper()]
#   polish = actual_polish_dict[polish.strip().upper()]
#   sym = actual_sym_dict[sym.strip().upper()]
#   return dict_key(weight, shape, color, clarity, fluor, cut, polish, sym)