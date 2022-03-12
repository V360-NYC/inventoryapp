import tempfile
import os
from google.cloud import storage
import numpy as np
import pandas as pd
import openpyxl
import uuid
import firebase_admin
from datetime import datetime 
from firebase_admin import credentials, firestore
import math
import time
import fnmatch
import pickle
from Mapping import *

tempdir = tempfile.gettempdir()
cred = credentials.ApplicationDefault()
firebase_admin.initialize_app(cred)
db = firestore.client()

tempdir = tempfile.mkdtemp()

client = storage.Client(project="friendlychat-bb9ff")

inventory_bucket = os.environ['INVENTORY_BUCKET']
master_bucket = os.environ['MASTER_BUCKET']
summary_bucket = os.environ['SUMMARY_BUCKET']

'''
    function to upload file to cloud storage bucket
    bucketName = bucket name where file is to be uploaded,
    path = path inside the bucket
    filepath = location of the file to be uploaded 

    assuming that file to be uploaded is always .xlsx file
'''

def uploadToBucket(bucketName, path, filepath):
    bucket = client.get_bucket(bucketName)

    blob = bucket.blob(path)
    
    with open(filepath, 'rb') as file:
        blob.upload_from_file(file)

    blob.make_public()



def downloadFromBucket(bucketName, path, filepath):
    bucket = client.get_bucket(bucketName)

    blob = bucket.blob(path)
    doesFileExist = blob.exists()
    print("download",filepath,path,bucketName)
    if not doesFileExist:
        raise Exception('remote file not present')
    
    blob.download_to_filename(filepath)

def exportDataFrameToExcel(dataframe, path):
    dataframe.to_excel(path, index = False)

def read_excel(filename):
    assert filename.split('.')[-1] in ['xlsx', 'xls'] ,'Not a excel file'
    return pd.read_excel(filename,engine='openpyxl')

def skip_to(df):
    for index, values in df.iterrows():
        for item in values:
            if item == 'REPORTNO':
                return pd.DataFrame(
                    df.iloc[index + 1 : ].to_numpy(), 
                    columns = df.iloc[index])
    raise Exception('header row not found')

def mapDate(row, CREATEDAT):
    if row['_merge']=='left_only':
        # old diamond ==> no change
        return row['CREATEDAT']
    elif row['_merge']=='right_only':
        return CREATEDAT

        # new diamond => first time appear
    else:
        return row['CREATEDAT']
    
def mapDate2(row, CREATEDAT):
    if row['_merge']=='left_only':
        # old diamond ==> no change
        return row['LASTMODIFIEDAT']
    elif row['_merge']=='right_only':
        return CREATEDAT

        # new diamond => first time appear
    else:
        return CREATEDAT     
        # updated

def map(row):
    if math.isnan(row['TOTAL_x']):
        return 'New'
    if row['TOTAL_x'] == row['TOTAL_y']:
        return 'Same'
    if row['TOTAL_x'] < row['TOTAL_y']:
        return 'Increase'
    if row['TOTAL_x'] > row['TOTAL_y']:
        return 'Decrease'

    # print(row['Total_x'] ,row['Total_y'])
    return 'Error'
    
def mapDiff(row):
    if math.isnan(row['TOTAL_x']):
        return 100
    if math.isnan(row['TOTAL_y']):
        return 'Error'
    if row['TOTAL_x'] == row['TOTAL_y']:
        return 0
    if row['TOTAL_x'] < row['TOTAL_y']:
        return (row['Total_y']-row['TOTAL_x'])*100/row['TOTAL_y']
    if row['TOTAL_x'] > row['TOTAL_y']:
        return (row['Total_y']-row['TOTAL_x'])*100/row['TOTAL_y']

    # print(row['Total_x'] ,row['Total_y'])
    return 'Error'
    
# def generate_summary_file(masterDate):
   
#     return summary

#     print(2)
    # summary = master.merge(inventory, how='right',on='ReportNo', indicator=True)
    # summary['summary'] = summary.apply(lambda row : map(row), axis = 1)
    # summary = summary[[column for column in summary.columns if '_y' in column] + ['ReportNo', 'summary']]
    # summary = summary.rename(columns={column:column.replace("_y", "") for column in summary.columns})
    
    # return summary

def generate_master_file(master, inventory):
    return pd.concat([master, inventory]).drop_duplicates(subset=['ReportNo'], keep = 'last')

def get_master_file_path(uid):
    collection_ref = db.collection('/'.join(['userFiles', uid, 'masterFiles'])).order_by('CREATEDAT', direction=firestore.Query.DESCENDING)

    docs = collection_ref.get()
    print("docs",docs)
    if len(docs) > 0:
        print(docs[0].to_dict()['filePath'])
        return docs[0].to_dict()['filePath']
      
    raise Exception('remote master not found')
  
def addSummaryFileMeta(summaryFilePath, uid, VENDORNAME):
  collection_ref = db.collection('/'.join(['userFiles', uid, 'summaryFiles']))

  data = {
    'filePath' : summaryFilePath,
    'bucket' : summary_bucket,
    'CREATEDAT' :datetime.now(),
    'downloadURL' : '/'.join(['https://storage.googleapis.com',summary_bucket, summaryFilePath]),
    'ack' : False,
    'VENDORNAME' : VENDORNAME
  }

  collection_ref.add(data)


def addMasterFileMeta(newMasterFilePath, uid, VENDORNAME):
  collection_ref = db.collection('/'.join(['userFiles', uid, 'masterFiles']))

  data = {
    'filePath' : newMasterFilePath,
    'bucket' : master_bucket,
    'CREATEDAT' : datetime.now(),
    'downloadURL' : '/'.join(['https://storage.googleapis.com',master_bucket, newMasterFilePath]),
    'ack' : False,
    'VENDORNAME' : VENDORNAME
  }

  collection_ref.add(data)

def mapVendor(row, VENDORNAME):
    if row['_merge']=='both':
        return VENDORNAME
    elif row['_merge']=='left_only':
        return row['VENDORNAME_x']
    else:
        return VENDORNAME  


def onInventoryFileUpload(inventory, uid, VENDORNAME, CREATEDAT):
    inventory['VENDORNAME'] = VENDORNAME

    start_time = time.time()
    try:
        
        masterFilePath = get_master_file_path(uid);
        downloadFromBucket(
            master_bucket,
            masterFilePath,
            os.path.join(tempdir, 'master.xlsx')
        )
        
    except Exception:
        
        masterColumns = {}
        print('exceptiong new dataframe columns',inventory.columns)

        for col in inventory.columns:
            masterColumns[col] = []
            

        for col in ['CREATEDAT', 'LASTMODIFIEDAT']:
            masterColumns[col] = []

        emptyMasterFile = pd.DataFrame(columns = masterColumns)
        print("new dataframe created")
        exportDataFrameToExcel(emptyMasterFile, os.path.join(tempdir, 'master.xlsx'))
    print(3)
    
    master = read_excel(os.path.join(tempdir,'master.xlsx'))
    print(3.1)
    master = mapping(os.path.join(tempdir, 'master.xlsx'))
    print(3.2)
    print("master columns",master.columns)
    master['REPORTNO'] = master['REPORTNO'].astype(np.int64)
    inventory['REPORTNO'] = inventory['REPORTNO'].astype(np.int64)
    print(4)
    masterDate= master.merge(inventory,how='outer',on='REPORTNO',indicator=True)
    print(masterDate.columns)
    print(5)
    masterDate['CREATEDAT'] = masterDate['CREATEDAT'].astype(str)
    masterDate['LASTMODIFIEDAT'] = masterDate['LASTMODIFIEDAT'].astype(str)
    masterDate['CREATEDAT'] = masterDate.apply(lambda row : mapDate(row, CREATEDAT), axis = 1)
    print(masterDate)
    print(5.1)
    masterDate['LASTMODIFIEDAT'] = masterDate.apply(lambda row : mapDate2(row, CREATEDAT), axis = 1)
    print(5.2)
    
    print(5.3)

    summary = masterDate[masterDate['_merge'] != 'left_only' ]

    summary['summary'] = summary.apply(lambda row : map(row), axis = 1)
    summary['%Change']= summary.apply(lambda row: mapDiff(row),axis=1)
    
    summary = summary[[column for column in summary.columns if '_y' in column] + ['REPORTNO','CREATEDAT','%Change','summary','LASTMODIFIEDAT']]
    summary=summary.rename(columns={column:column.replace("_y", "") for column in summary.columns})

    summary['REPORTNOabc']=summary['REPORTNO']
    summary['%Changeabc']=summary['%Change']
    summary['summaryabc']=summary['summary']

    summary=summary.drop(['REPORTNO','%Change','summary'], axis = 1)
    summary=summary[[column for column in summary.columns if 'abc' in column]+ [column for column in summary.columns if 'abc' not in column]]
    summary=summary.rename(columns={column:column.replace("abc", "") for column in summary.columns})

    print(summary.columns)
    
    masterDate=masterDate[[column for column in masterDate.columns if '_y' not in column]] 
    masterDate=masterDate.rename(columns={column:column.replace("_x", "") for column in masterDate.columns})
    print(masterDate.columns)

    masterDate=pd.concat([masterDate,summary]).drop_duplicates(subset=['REPORTNO'],keep='last')
    masterDate=masterDate.drop(['summary','_merge','%Change'],axis=1)

    
    print(6)
    masterDate['REPORTNOabc']=masterDate['REPORTNO']

    # masterDate['VIDEOLINKabc']=masterDate['VIDEOLINK']
    # masterDate=masterDate.drop(['REPORTNO','VIDEOLINK'], axis = 1)
    masterDate=masterDate.drop(['REPORTNO'], axis = 1)
    masterDate=masterDate[[column for column in masterDate.columns if 'abc' in column]+ [column for column in masterDate.columns if 'abc' not in column]]
    masterDate=masterDate.rename(columns={column:column.replace("abc", "") for column in masterDate.columns})
    
    print(8)
    start_time = time.time()
    exportDataFrameToExcel(summary, os.path.join(tempdir, 'summary.xlsx'))
    summaryFilePath = '/'.join([uid, str(uuid.uuid4()), 'summary.xlsx'])

    uploadToBucket(
        summary_bucket,
        summaryFilePath,
        os.path.join(tempdir, 'summary.xlsx')
    )
    
    addSummaryFileMeta(summaryFilePath, uid, VENDORNAME)
    print(9)
    start_time = time.time()

    exportDataFrameToExcel(masterDate, os.path.join(tempdir, 'master.xlsx'))

    folder =  str(uuid.uuid4())
    newMasterFilePath = '/'.join([uid, folder, 'master.xlsx'])

    uploadToBucket(
        master_bucket,
        newMasterFilePath,
        os.path.join(tempdir, 'master.xlsx')
    )

    with open(os.path.join(tempdir, 'master.pickle'), 'wb+') as pickleFile:
        pickle.dump(masterDate, pickleFile)

    uploadToBucket(
        master_bucket,
        '/'.join([uid, folder, 'master.pickle']),
        os.path.join(tempdir, 'master.pickle')
    )
    
    addMasterFileMeta(newMasterFilePath, uid, VENDORNAME)
    print(10)
    print("{} entire data analysis code took ".format(time.time() - start_time))


def hello_firestore(event, context):
    """
    Triggered by a change to a Firestore document.
    Args:
        event (dict): Event payload.
        context (google.cloud.functions.Context): Metadata for the event.
    """
    print(event)
    resource_string = context.resource
    bucketName= event['value']['fields']['bucket']['stringValue']
    bucket = client.get_bucket(bucketName)

    bucketPathArray = event['value']['fields']['files']['arrayValue']['values']
    filenames=[]

    userId = None
    
    for everyobj in bucketPathArray:
        currentFilePath=everyobj['mapValue']['fields']['filePath']['stringValue']
        print(currentFilePath)
        blob=bucket.blob(currentFilePath)
        blob.download_to_filename(os.path.join(tempdir, currentFilePath.split('/')[-1]))
        globaldf=mapping(os.path.join(tempdir, currentFilePath.split('/')[-1]))
        userId=currentFilePath.split('/')[0]
        filenames.append(os.path.join(tempdir,currentFilePath.split('/')[-1]))
        
    def findrowscount(ws):
        a=ws.values
        sum=0
        for obj in a:
            sum+=1
        return sum
    
    def findVideoColumn(ws):
        flag=1
        pos=-1
        for i in ws.values:
            
            count=1
            for value in i:
                if value=='VideoLink':
                        return flag,count
                if value=='REPORTNO':
                        pos=count
                        flag2=flag
                count+=1
            flag+=1
        if pos==-1:
            return -1,-1    
        return flag2,pos     

    def extractLinks(VideoColumn, sumRows,HeaderRow):
        videoLinkList=[]
        for i in range(HeaderRow+1, sumRows+1):
            try:
                videoLinkList.append(ws.cell(row=i,column=VideoColumn).hyperlink.target)

            except:

                videoLinkList.append('None')
        return videoLinkList

    def makeDFwithVidLink(ws):
        sumRows=findrowscount(ws)
        headerIndex, VideoColumn = findVideoColumn(ws)
        if headerIndex == -1:
            return -1
        df=pd.DataFrame(ws.values)
        df=df[~(df.index<headerIndex-1)]
        df.columns=df.iloc[0]
        df=df.iloc[1:]
        if VideoColumn !=-1:
            videoLinks= extractLinks(VideoColumn,sumRows,headerIndex)        
            df['VideoFullLinks']=videoLinks
        else :
            df['VideoFullLinks'] = 'None'

        # df=df[ ~df['REPORTNO'].isnull()]
        # df['REPORTNO']=df['REPORTNO'].astype('int64')
        # df['Total'] = df['Total'].astype('int64')
        # df['Weight'] = df['Weight'].astype(float)
        # df['Rap Total'] = df['Rap Total'].astype(float)

        return df

    # flagval=0
    # globaldf=pd.DataFrame()
    # errorList=[]
    # for name in filenames:
    #     wb = openpyxl.load_workbook(name,data_only=True)
    #     sheetName=wb.sheetnames
    #     for iterator in sheetName:
    #         ws = wb[iterator]
    #         df=makeDFwithVidLink(ws)
    #         if type(df)==int :
    #             errorList.append(f'The sheet named {iterator} in the file {file} had a missing ReportNo column and hence could not be processed')
    #             print('Not a valid file')
    #             print(iterator , name,'gave error')
    #         else :
    #             if flagval==0:
    #                 globaldf=globaldf.append(df)
    #                 flagval=1
    #             else:
    #                 globaldf=pd.concat([globaldf,df]).drop_duplicates(subset=['REPORTNO'],keep='last')
    


    def exportDataFrameToExcel(dataframe, path):
        dataframe.to_excel(path)

    exportDataFrameToExcel(globaldf, os.path.join(tempdir, 'inventory.xlsx'))
    
    def uploadToBucket(bucketName, path, filepath):
        bucket = client.get_bucket(bucketName)

        blob = bucket.blob(path)
        
        with open(filepath, 'rb') as file:
            blob.upload_from_file(file)

        blob.make_public()
        
    tempuuid=uuid.uuid4()
    tempuuid=str(tempuuid)
    
    assert userId != None

    uploadToBucket('business-inventory-files', '/'.join([userId,tempuuid,'inventory.xlsx']),os.path.join(tempdir,'inventory.xlsx'))
    metaData={}
    metaData['bucket']='business-inventory-files'
    metaData['CREATEDAT']=datetime.fromtimestamp(int(event['value']['fields']['CREATEDAT']['integerValue'])/1000.0)
    print('created at', metaData['CREATEDAT'], type(metaData['CREATEDAT']))
    metaData['downloadURL']='https://storage.googleapis.com/business-inventory-files/'+userId+'/'+tempuuid+'/inventory.xlsx'
    # metaData['errorList']=errorList
    metaData['filePath']=userId+'/'+tempuuid+'/inventory.xlsx'
    metaData['VENDORNAME'] = event['value']['fields']['VENDORNAME']['stringValue']
    doc_ref=db.collection('userFiles').document(userId).collection('inventoryFiles').add(metaData)
    
    # start data analysis

    CREATEDAT = metaData['CREATEDAT'].strftime("%d/%m/%Y")
    print('created at 2', CREATEDAT, type(CREATEDAT))
    print("metaData ",metaData)
    print("d",doc_ref)
    
    onInventoryFileUpload(globaldf, userId, metaData['VENDORNAME'], CREATEDAT)



